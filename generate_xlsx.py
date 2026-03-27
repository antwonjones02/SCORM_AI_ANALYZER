#!/usr/bin/env python3
"""
SCORM AI Readiness Report — XLSX Generator by Claw 🦊
Turns batch_report.json into a multi-tab Excel workbook.

Usage:
    python3 generate_xlsx.py [--input batch_report.json] [--output report.xlsx] [--client "Acme Corp"]

Tabs:
    1. Executive Summary   — KPIs, tier breakdown, catalog health at a glance
    2. Course Catalog      — Per-course scores, flags, skills (the working spreadsheet)
    3. Remediation Plan    — Prioritized action list, sortable by score/tier
    4. Skills Inventory    — All inferred skills with frequency counts
    5. Quality Flags       — All flag types with affected course counts
"""

import json
import argparse
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("❌ openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Colors ────────────────────────────────────────────────────────────────────
NAVY    = "003087"
RED     = "C8102E"
GREEN   = "28a745"
ORANGE  = "FF8C00"
BLUE    = "1a8fc1"
WHITE   = "FFFFFF"
LGRAY   = "F5F5F5"
MGRAY   = "E0E0E0"
DGRAY   = "2D2D2D"
YELLOW  = "FFF9C4"

# ── Style helpers ─────────────────────────────────────────────────────────────
def hfont(bold=True, color=WHITE, size=11):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def bfont(bold=False, color=DGRAY, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color=MGRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color=NAVY):
    return Border(bottom=Side(style="medium", color=color))

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def tier_color(score):
    if score is None: return MGRAY
    if score >= 91: return GREEN
    if score >= 71: return BLUE
    if score >= 41: return ORANGE
    return RED

def tier_label(score):
    if score is None: return "Unknown"
    if score >= 91: return "AI Ready"
    if score >= 71: return "Good"
    if score >= 41: return "Needs Work"
    return "Not Ready"

def score_bar_text(score):
    """Text-based score bar for cells."""
    if score is None: return "N/A"
    filled = int(score / 10)
    return f"{'█' * filled}{'░' * (10 - filled)}  {score}/100"

def header_row(ws, row, cols, bg=NAVY, fg=WHITE, size=10):
    """Write a styled header row."""
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, color=fg, size=size)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()

def data_row(ws, row, values, alt=False, bold_col=None):
    """Write a data row with alternating shading."""
    bg = LGRAY if alt else WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.alignment = left()
        c.border = thin_border()
        if bold_col and col == bold_col:
            c.font = Font(name="Calibri", bold=True, size=10, color=DGRAY)
        else:
            c.font = bfont()

def color_score_cell(ws, row, col, score):
    """Color a score cell based on tier."""
    c = ws.cell(row=row, column=col)
    tc = tier_color(score)
    c.fill = fill(tc)
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    c.alignment = center()

def color_tier_cell(ws, row, col, score):
    """Color a tier label cell based on score."""
    c = ws.cell(row=row, column=col)
    tc = tier_color(score)
    c.fill = fill(tc)
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    c.alignment = center()


# ── TAB 1: Executive Summary ──────────────────────────────────────────────────
def build_summary(wb, data, client_name):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    courses = data.get("courses", [])
    total = data.get("total_packages", 0)
    processed = data.get("processed", 0)
    avg = data.get("average_ai_readiness_score")
    dist = data.get("score_distribution", {})
    generated = data.get("generated_at", "")[:10]
    llm = data.get("llm_enriched", False)

    ai_ready = dist.get("AI Ready", 0)
    good = dist.get("Good", 0)
    needs_work = dist.get("Needs Work", 0)
    not_ready = dist.get("Not Ready", 0)
    unknown = dist.get("Unknown", 0)

    ready_pct = round(100 * (ai_ready + good) / processed, 1) if processed else 0
    remediation_pct = round(100 * (needs_work + not_ready) / processed, 1) if processed else 0

    # Title block
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "LMS AI READINESS ASSESSMENT"
    c.font = Font(name="Calibri", bold=True, size=18, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"{client_name}  |  Assessment Date: {generated}  |  {total} Packages Analyzed  |  AI Enrichment: {'ON' if llm else 'OFF'}"
    c.font = Font(name="Calibri", size=10, color=WHITE, italic=True)
    c.fill = fill(RED)
    c.alignment = center()
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 12

    # KPI cards (row 4-8)
    kpis = [
        ("TOTAL COURSES", str(total), NAVY),
        ("AVG READINESS SCORE", f"{avg}/100" if avg else "N/A", RED if (avg or 0) < 41 else ORANGE if (avg or 0) < 71 else GREEN),
        ("AI READY + GOOD", f"{ai_ready + good} courses ({ready_pct}%)", GREEN),
        ("NEEDS REMEDIATION", f"{needs_work + not_ready} courses ({remediation_pct}%)", RED),
    ]

    for i, (label, value, color) in enumerate(kpis):
        col = i * 2 + 1
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        ws.merge_cells(start_row=5, start_column=col, end_row=7, end_column=col+1)
        ws.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)

        lc = ws.cell(row=4, column=col, value=label)
        lc.font = Font(name="Calibri", bold=True, size=8, color=color)
        lc.fill = fill(LGRAY)
        lc.alignment = center()
        ws.row_dimensions[4].height = 16

        vc = ws.cell(row=5, column=col, value=value)
        vc.font = Font(name="Calibri", bold=True, size=20, color=color)
        vc.fill = fill(WHITE)
        vc.alignment = center()
        for r in range(5, 8):
            ws.row_dimensions[r].height = 22

        bc = ws.cell(row=8, column=col)
        bc.fill = fill(color)
        ws.row_dimensions[8].height = 6

    ws.row_dimensions[9].height = 12

    # Score Distribution table
    ws.merge_cells("A10:H10")
    c = ws["A10"]
    c.value = "SCORE DISTRIBUTION"
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[10].height = 22

    header_row(ws, 11, ["Tier", "Score Range", "Count", "% of Catalog", "What It Means", "Action Required", "", ""], bg=DGRAY)

    tiers = [
        ("AI Ready",   "91–100", ai_ready,   GREEN,   "Excellent metadata. AI personalization ready.",   "Monitor & maintain"),
        ("Good",       "71–90",  good,        BLUE,    "Minor gaps. Can power basic recommendations.",    "Minor enrichment"),
        ("Needs Work", "41–70",  needs_work,  ORANGE,  "Significant cleanup needed before AI use.",       "Priority remediation"),
        ("Not Ready",  "0–40",   not_ready,   RED,     "Major metadata gaps. AI will fail.",              "Immediate remediation or retire"),
    ]
    if unknown:
        tiers.append(("Unknown", "N/A", unknown, MGRAY, "Not enriched with AI analysis.", "Run --llm mode"))

    for i, (tier, rng, count, color, meaning, action) in enumerate(tiers):
        row = 12 + i
        pct = f"{round(100*count/processed,1)}%" if processed else "0%"
        data_row(ws, row, [tier, rng, count, pct, meaning, action, "", ""], alt=(i%2==0))
        ws.cell(row=row, column=1).fill = fill(color)
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        ws.cell(row=row, column=3).alignment = center()
        ws.cell(row=row, column=4).alignment = center()
        ws.row_dimensions[row].height = 22

    # Spacer
    ws.row_dimensions[17].height = 12

    # Key Findings
    ws.merge_cells("A18:H18")
    c = ws["A18"]
    c.value = "KEY FINDINGS"
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[18].height = 22

    # Count metadata stats
    no_desc = sum(1 for c in courses if not c.get("has_description"))
    no_kw = sum(1 for c in courses if not c.get("has_keywords"))
    no_skills = sum(1 for c in courses if not c.get("inferred_skills"))
    has_flags = sum(1 for c in courses if c.get("quality_flags"))

    findings = [
        ("Courses missing description", no_desc, f"{round(100*no_desc/processed,0):.0f}%" if processed else "0%", RED if no_desc > processed * 0.5 else ORANGE),
        ("Courses missing keywords", no_kw, f"{round(100*no_kw/processed,0):.0f}%" if processed else "0%", RED if no_kw > processed * 0.5 else ORANGE),
        ("Courses with no skills inferred", no_skills, f"{round(100*no_skills/processed,0):.0f}%" if processed else "0%", ORANGE),
        ("Courses with quality flags", has_flags, f"{round(100*has_flags/processed,0):.0f}%" if processed else "0%", ORANGE),
    ]

    header_row(ws, 19, ["Finding", "Count", "% of Catalog", "Status", "", "", "", ""], bg=DGRAY)
    for i, (label, count, pct, color) in enumerate(findings):
        row = 20 + i
        data_row(ws, row, [label, count, pct, tier_label(100 - (count/max(processed,1))*100), "", "", "", ""], alt=(i%2==0))
        ws.cell(row=row, column=4).fill = fill(color)
        ws.cell(row=row, column=4).font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        ws.cell(row=row, column=4).alignment = center()
        ws.row_dimensions[row].height = 20

    # Column widths
    widths = [20, 12, 10, 14, 38, 28, 5, 5]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "A3"


# ── TAB 2: Course Catalog ─────────────────────────────────────────────────────
def build_catalog(wb, data):
    ws = wb.create_sheet("Course Catalog")
    ws.sheet_view.showGridLines = False

    courses = data.get("courses", [])
    llm = data.get("llm_enriched", False)

    # Headers
    cols = [
        "File", "Course Title", "SCORM Ver", "Modules", "SCOs",
        "Has Description", "Has Keywords", "Keyword Count",
        "AI Score", "Tier", "Difficulty", "Target Audience",
        "Inferred Skills", "Quality Flags", "AI Readiness Notes"
    ]
    header_row(ws, 1, cols, bg=NAVY)
    ws.row_dimensions[1].height = 28

    for i, course in enumerate(courses):
        row = i + 2
        score = course.get("ai_readiness_score")
        skills = ", ".join(course.get("inferred_skills", [])) or "—"
        flags = ", ".join(course.get("quality_flags", [])) or "—"

        values = [
            course.get("file", ""),
            course.get("title", "Untitled") or course.get("file", ""),
            course.get("scorm_version", "—"),
            course.get("total_modules", 0),
            course.get("sco_count", 0),
            "Yes" if course.get("has_description") else "No",
            "Yes" if course.get("has_keywords") else "No",
            course.get("keyword_count", 0),
            score if score is not None else "N/A",
            tier_label(score),
            course.get("difficulty_level", "—") or "—",
            course.get("target_audience", "—") or "—",
            skills,
            flags,
            course.get("ai_readiness_notes", "—") or "—",
        ]
        data_row(ws, row, values, alt=(i%2==0), bold_col=2)

        # Color score + tier cells
        if score is not None:
            color_score_cell(ws, row, 9, score)
            ws.cell(row=row, column=9).value = score
            color_tier_cell(ws, row, 10, score)
            ws.cell(row=row, column=10).value = tier_label(score)

        # Color Yes/No cells
        for col, field in [(6, "has_description"), (7, "has_keywords")]:
            c = ws.cell(row=row, column=col)
            if course.get(field):
                c.fill = fill("E8F5E9")
                c.font = Font(name="Calibri", color=GREEN, bold=True, size=10)
            else:
                c.fill = fill("FFEBEE")
                c.font = Font(name="Calibri", color=RED, bold=True, size=10)
            c.alignment = center()

        ws.row_dimensions[row].height = 40

    # Column widths
    widths = [28, 40, 10, 10, 8, 14, 14, 12, 12, 14, 16, 28, 40, 40, 40]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


# ── TAB 3: Remediation Plan ───────────────────────────────────────────────────
def build_remediation(wb, data):
    ws = wb.create_sheet("Remediation Plan")
    ws.sheet_view.showGridLines = False

    courses = data.get("courses", [])
    sorted_courses = sorted(courses, key=lambda c: (c.get("ai_readiness_score") or 0))

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "REMEDIATION PLAN — Sorted by Priority (Lowest Score First)"
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    cols = [
        "Priority", "Course Title", "AI Score", "Tier",
        "Phase", "Primary Action", "Est. Score Lift",
        "Has Desc", "Has Keywords", "Quality Flags"
    ]
    header_row(ws, 2, cols, bg=DGRAY)
    ws.row_dimensions[2].height = 24

    def get_phase(score):
        if score is None: return "Phase 1"
        if score < 41: return "Phase 1 — Immediate"
        if score < 71: return "Phase 2 — 30-60 Days"
        if score < 91: return "Phase 3 — 60-90 Days"
        return "Phase 4 — Maintain"

    def get_action(course):
        score = course.get("ai_readiness_score")
        flags = course.get("quality_flags", [])
        if not course.get("has_description"): return "Add course description (75+ words)"
        if not course.get("has_keywords"): return "Add keyword tags (5-10 terms)"
        if not course.get("inferred_skills"): return "Map to skills taxonomy"
        if score and score < 41: return "Full metadata enrichment required"
        if score and score < 71: return "Enrich description + add skills tags"
        return "Minor polish — verify difficulty + audience"

    def get_lift(score):
        if score is None: return "+20-30 pts (est.)"
        if score < 41: return "+40-55 pts (est.)"
        if score < 71: return "+20-30 pts (est.)"
        if score < 91: return "+5-10 pts (est.)"
        return "Maintain"

    for i, course in enumerate(sorted_courses):
        row = i + 3
        score = course.get("ai_readiness_score")
        flags = ", ".join(course.get("quality_flags", [])) or "—"

        values = [
            i + 1,
            course.get("title", "") or course.get("file", ""),
            score if score is not None else "N/A",
            tier_label(score),
            get_phase(score),
            get_action(course),
            get_lift(score),
            "Yes" if course.get("has_description") else "No",
            "Yes" if course.get("has_keywords") else "No",
            flags,
        ]
        data_row(ws, row, values, alt=(i%2==0), bold_col=2)

        if score is not None:
            color_score_cell(ws, row, 3, score)
            ws.cell(row=row, column=3).value = score
            color_tier_cell(ws, row, 4, score)
            ws.cell(row=row, column=4).value = tier_label(score)

        for col, field in [(8, "has_description"), (9, "has_keywords")]:
            c = ws.cell(row=row, column=col)
            if course.get(field):
                c.fill = fill("E8F5E9"); c.font = Font(name="Calibri", color=GREEN, bold=True, size=10)
            else:
                c.fill = fill("FFEBEE"); c.font = Font(name="Calibri", color=RED, bold=True, size=10)
            c.alignment = center()

        ws.row_dimensions[row].height = 36

    widths = [10, 40, 12, 14, 24, 42, 18, 10, 12, 40]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"


# ── TAB 4: Skills Inventory ───────────────────────────────────────────────────
def build_skills(wb, data):
    ws = wb.create_sheet("Skills Inventory")
    ws.sheet_view.showGridLines = False

    courses = data.get("courses", [])

    # Aggregate skills
    skill_map = {}
    for course in courses:
        for skill in course.get("inferred_skills", []):
            if skill not in skill_map:
                skill_map[skill] = {"count": 0, "courses": []}
            skill_map[skill]["count"] += 1
            skill_map[skill]["courses"].append(course.get("title", course.get("file", "")))

    sorted_skills = sorted(skill_map.items(), key=lambda x: -x[1]["count"])

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "SKILLS INVENTORY — All inferred skills from catalog AI analysis"
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    header_row(ws, 2, ["Rank", "Skill", "Course Count", "Frequency", "Courses Teaching This Skill"], bg=DGRAY)
    ws.row_dimensions[2].height = 24

    total_courses = max(len(courses), 1)
    for i, (skill, info) in enumerate(sorted_skills):
        row = i + 3
        count = info["count"]
        freq = f"{round(100*count/total_courses,1)}%"
        course_list = " | ".join(info["courses"][:5])
        if len(info["courses"]) > 5:
            course_list += f" (+{len(info['courses'])-5} more)"

        data_row(ws, row, [i+1, skill, count, freq, course_list], alt=(i%2==0))
        ws.cell(row=row, column=1).alignment = center()
        ws.cell(row=row, column=3).alignment = center()
        ws.cell(row=row, column=4).alignment = center()

        # Color by frequency
        if count >= 3:
            ws.cell(row=row, column=2).fill = fill("E8F5E9")
        elif count == 1:
            ws.cell(row=row, column=2).fill = fill("FFF9C4")

        ws.row_dimensions[row].height = 20

    if not sorted_skills:
        ws.cell(row=3, column=1).value = "No skills data available — run with --llm flag to enable AI enrichment"
        ws.cell(row=3, column=1).font = Font(italic=True, color="999999")

    widths = [8, 40, 14, 12, 60]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)


# ── TAB 5: Quality Flags ──────────────────────────────────────────────────────
def build_flags(wb, data):
    ws = wb.create_sheet("Quality Flags")
    ws.sheet_view.showGridLines = False

    courses = data.get("courses", [])

    # Aggregate flags
    flag_map = {}
    for course in courses:
        for flag in course.get("quality_flags", []):
            if flag not in flag_map:
                flag_map[flag] = {"count": 0, "courses": []}
            flag_map[flag]["count"] += 1
            flag_map[flag]["courses"].append(course.get("title", course.get("file", "")))

    sorted_flags = sorted(flag_map.items(), key=lambda x: -x[1]["count"])

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "QUALITY FLAGS — Issues detected across catalog (sorted by frequency)"
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = fill(RED)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    header_row(ws, 2, ["Rank", "Quality Flag", "Courses Affected", "% of Catalog", "Affected Courses"], bg=DGRAY)
    ws.row_dimensions[2].height = 24

    total = max(len(courses), 1)
    for i, (flag, info) in enumerate(sorted_flags):
        row = i + 3
        count = info["count"]
        pct = f"{round(100*count/total,1)}%"
        course_list = " | ".join(info["courses"][:5])
        if len(info["courses"]) > 5:
            course_list += f" (+{len(info['courses'])-5} more)"

        data_row(ws, row, [i+1, flag, count, pct, course_list], alt=(i%2==0))
        ws.cell(row=row, column=1).alignment = center()
        ws.cell(row=row, column=3).alignment = center()
        ws.cell(row=row, column=4).alignment = center()

        # Color severity
        severity = fill("FFEBEE") if count >= total * 0.5 else fill("FFF9C4") if count >= total * 0.25 else fill("FFFFFF")
        ws.cell(row=row, column=2).fill = severity
        ws.row_dimensions[row].height = 20

    if not sorted_flags:
        ws.cell(row=3, column=1).value = "No quality flags detected — run with --llm flag to enable AI enrichment"
        ws.cell(row=3, column=1).font = Font(italic=True, color="999999")

    widths = [8, 50, 16, 14, 60]
    for i, w in enumerate(widths, 1):
        set_col_width(ws, i, w)


# ── Main ──────────────────────────────────────────────────────────────────────
def generate_xlsx(input_file, output_file, client_name):
    data = json.loads(Path(input_file).read_text())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    print(f"📊 Building XLSX for: {client_name}")
    print(f"   Courses: {data.get('processed', 0)}")

    build_summary(wb, data, client_name)
    print("   ✅ Tab 1: Executive Summary")

    build_catalog(wb, data)
    print("   ✅ Tab 2: Course Catalog")

    build_remediation(wb, data)
    print("   ✅ Tab 3: Remediation Plan")

    build_skills(wb, data)
    print("   ✅ Tab 4: Skills Inventory")

    build_flags(wb, data)
    print("   ✅ Tab 5: Quality Flags")

    wb.save(output_file)
    size_kb = Path(output_file).stat().st_size // 1024
    print(f"\n✅ XLSX saved: {output_file} ({size_kb} KB)")


def main():
    parser = argparse.ArgumentParser(description="Generate XLSX report from SCORM batch analysis")
    parser.add_argument("--input", default="batch_report.json", help="Input JSON file")
    parser.add_argument("--output", default="scorm_ai_readiness_report.xlsx", help="Output XLSX file")
    parser.add_argument("--client", default="Client Organization", help="Client name")
    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"❌ Input not found: {args.input}")
        sys.exit(1)

    generate_xlsx(args.input, args.output, args.client)


if __name__ == "__main__":
    main()
